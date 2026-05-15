import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Cột xuất ra Excel: (field_key, tên hiển thị, độ rộng cột)
COLUMNS = [
    ('so_hieu',          'Số hiệu',           22),
    ('ten_van_ban',      'Tên văn bản',        55),
    ('loai_van_ban',     'Loại văn bản',       20),
    ('co_quan_ban_hanh', 'Cơ quan ban hành',   32),
    ('nguoi_ky',         'Người ký',           25),
    ('ngay_ban_hanh',    'Ngày ban hành',      15),
    ('ngay_hieu_luc',    'Ngày hiệu lực',      15),
    ('trich_yeu',        'Trích yếu',          60),
    ('noi_dung',         'Nội dung toàn văn',  80),
    ('file_dinh_kem',    'File đính kèm',      50),
    ('url_goc',          'URL gốc',            55),
    ('nguon',            'Nguồn',              30),
]

HEADER_BG   = '1F4E79'
HEADER_FG   = 'FFFFFF'
ALT_ROW_BG  = 'EBF3FA'
THIN_BORDER = Border(
    bottom=Side(style='thin', color='CCCCCC'),
)


def save_to_excel(docs_by_source: dict, output_dir: str = 'output') -> str:
    """
    Lưu kết quả vào file Excel nhiều sheet.
    docs_by_source: dict { tên_sheet: [list doc dicts] }
    Trả về đường dẫn file đã lưu.
    """
    Path(output_dir).mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filepath = Path(output_dir) / f'van_ban_giao_duc_{timestamp}.xlsx'

    wb = openpyxl.Workbook()

    # Xóa sheet mặc định
    default = wb.active
    wb.remove(default)

    all_docs = []
    for sheet_name, docs in docs_by_source.items():
        if not docs:
            continue
        all_docs.extend(docs)
        ws = wb.create_sheet(title=_safe_sheet_name(sheet_name))
        _write_sheet(ws, docs)

    # Sheet "Tổng hợp" nếu có nhiều nguồn
    if len([v for v in docs_by_source.values() if v]) > 1 and all_docs:
        ws_all = wb.create_sheet(title='Tổng hợp', index=0)
        _write_sheet(ws_all, all_docs)

    wb.save(filepath)
    logger.info(f"Đã lưu {len(all_docs)} văn bản → {filepath}")
    return str(filepath)


def _safe_sheet_name(name: str) -> str:
    """Xóa ký tự không hợp lệ trong tên sheet Excel (max 31 chars)"""
    invalid = r'[:\\/?*\[\]]'
    import re
    cleaned = re.sub(invalid, '', name).strip()
    return cleaned[:31] or 'Sheet'


def _write_sheet(ws, docs: list):
    """Ghi dữ liệu vào worksheet với định dạng đẹp"""

    # --- Header ---
    header_font   = Font(name='Calibri', bold=True, color=HEADER_FG, size=11)
    header_fill   = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type='solid')
    header_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    alt_fill      = PatternFill(start_color=ALT_ROW_BG, end_color=ALT_ROW_BG, fill_type='solid')
    data_align    = Alignment(vertical='top', wrap_text=True)

    for col_idx, (_, label, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        ws.column_dimensions[get_column_letter(col_idx)].width = min(width, 60)

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = 'A2'

    # --- Dữ liệu ---
    for row_idx, doc in enumerate(docs, 2):
        for col_idx, (field, _, _) in enumerate(COLUMNS, 1):
            value = str(doc.get(field, '') or '')
            # Giới hạn nội dung dài để tránh lỗi Excel (max 32767 ký tự/cell)
            if field == 'noi_dung' and len(value) > 30000:
                value = value[:30000] + '\n... [Xem link gốc để đọc đầy đủ]'
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = data_align
            cell.border    = THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Auto filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
